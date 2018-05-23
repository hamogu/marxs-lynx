import argparse

from astropy.table import Table
import astropy.units as u
import openpyxl

parser = argparse.ArgumentParser(description="Convert xlsx file from Ralf into ecsv files for my simulatons. This assumes that Ralf's tables adheres to his usual format.")
parser.add_argument('filename',
                    help='xlsx table from Ralf')
parser.add_argument('outfile',
                    help='filename and path to ecsv file to be written')
parser.add_argument('--comment', '-c', default='',
                    help='comment for file header (e.g. email data when file was received etc.)')


args = parser.parse_args()


class InconsistentTableError(Exception):
    pass

wb = openpyxl.load_workbook(args.filename)
ws = wb.active

# Check that this matches Ralf's usual format:
if ((ws['A3'].value != "lambda") |
    (ws['B3'].value != "angle") |
    (ws['A4'].value != "[nm]") |
    (ws['B4'].value != "[deg]")):
    raise InconsistentTableError("Format does not match Ralf's usual xlsx files")

tab = Table(names=['lambda', 'theta'] +
            ['o{}'.format(ws.cell(row=4, column=i).value) for i in range(3, ws.max_column + 1)],
            dtype=['f4'] * (ws.max_column))
for row in ws.iter_rows(min_row=5):
    tab.add_row([cell.value for cell in row])

tab['lambda'].unit = u.Angstrom
tab['theta'].unit = u.degree
tab.meta['ORIGFILE'] = args.filename
tab.meta['author'] = 'converted by Hans Moritz Guenther'
tab.meta['origin'] = 'Ralf Heilmann'
tab.meta['description'] = [cell[0].value for cell in ws['A1': 'H1']]
tab.meta['comment'] = args.comment

tab.write(args.outfile, format='ascii.ecsv', overwrite=True)
